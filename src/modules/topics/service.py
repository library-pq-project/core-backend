import csv
import io
import json

from fastapi import HTTPException, status
from sqlalchemy.exc import IntegrityError

from src.common.utils import generate_slug
from src.modules.topics.models import Topic
from src.modules.topics.repository import TopicRepository
from src.modules.topics.schemas import (
    ImportRowError,
    TopicBulkRow,
    TopicBulkUpsertRequest,
    TopicBulkUpsertResult,
    TopicCreate,
    TopicUpdate,
)


class TopicService:
    def __init__(self, repository: TopicRepository):
        self.repository = repository

    def list_topics(self, *, course_id: int | None = None, skip: int, limit: int) -> list[Topic]:
        return self.repository.list(course_id=course_id, skip=skip, limit=limit)

    def get_topic(self, topic_id: int) -> Topic:
        topic = self.repository.get(topic_id)
        if not topic:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Topic not found")
        return topic

    def get_topic_by_slug(self, topic_slug: str) -> Topic:
        topic = self.repository.get_by_slug(topic_slug)
        if not topic:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Topic not found")
        return topic

    def list_topics_in_course_slug(self, course_slug: str, *, skip: int, limit: int) -> list[Topic]:
        return self.repository.list_by_course_slug(course_slug, skip=skip, limit=limit)

    def create_topic(self, payload: TopicCreate) -> Topic:
        topic = Topic(
            course_id=payload.course_id,
            name=payload.name,
            slug=generate_slug(payload.name),
            description=payload.description,
        )
        try:
            return self.repository.create(topic)
        except IntegrityError as exc:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Unable to create topic with provided values",
            ) from exc

    def update_topic(self, topic_id: int, payload: TopicUpdate) -> Topic:
        topic = self.get_topic(topic_id)
        updates = payload.model_dump(exclude_unset=True)

        if "course_id" in updates:
            topic.course_id = updates["course_id"]
        if "name" in updates:
            topic.name = updates["name"]
        if "name" in updates:
            topic.slug = generate_slug(updates["name"])
        if "description" in updates:
            topic.description = updates["description"]

        try:
            return self.repository.save(topic)
        except IntegrityError as exc:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Unable to update topic with provided values",
            ) from exc

    def delete_topic(self, topic_id: int) -> None:
        topic = self.get_topic(topic_id)
        self.repository.delete(topic)

    def _upsert_topic_row(self, row: TopicBulkRow) -> tuple[str, int]:
        name = row.name.strip()
        if not name:
            raise ValueError("name is required")
        slug = generate_slug(name)
        existing = self.repository.get_by_course_and_slug(course_id=row.course_id, slug=slug)
        if existing is None:
            existing_by_name = self.repository.get_by_course_and_name(course_id=row.course_id, name=name)
            if existing_by_name is not None:
                existing = existing_by_name

        if existing is None:
            created = self.repository.create(
                Topic(
                    course_id=row.course_id,
                    name=name,
                    slug=slug,
                    description=row.description,
                )
            )
            return "created", created.id

        existing.name = name
        existing.description = row.description
        if existing.slug != slug:
            existing.slug = slug
        saved = self.repository.save(existing)
        return "updated", saved.id

    def bulk_upsert_topics(self, payload: TopicBulkUpsertRequest) -> TopicBulkUpsertResult:
        created_count = 0
        updated_count = 0
        skipped_count = 0
        topic_ids: list[int] = []
        errors: list[ImportRowError] = []

        for index, row in enumerate(payload.rows, start=1):
            try:
                action, topic_id = self._upsert_topic_row(row)
                topic_ids.append(topic_id)
                if action == "created":
                    created_count += 1
                else:
                    updated_count += 1
            except Exception as exc:  # noqa: BLE001
                skipped_count += 1
                errors.append(ImportRowError(row_number=index, errors=[str(exc)]))

        return TopicBulkUpsertResult(
            created_count=created_count,
            updated_count=updated_count,
            skipped_count=skipped_count,
            errors=errors,
            topic_ids=sorted(set(topic_ids)),
        )

    def _parse_file_rows(self, *, file_name: str, content: bytes) -> list[dict]:
        extension = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""
        if extension == "json":
            loaded = json.loads(content.decode("utf-8"))
            if isinstance(loaded, dict):
                loaded = loaded.get("rows", [])
            if not isinstance(loaded, list):
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="JSON upload must contain a list of rows")
            return [dict(item) for item in loaded if isinstance(item, dict)]

        if extension == "csv":
            text = content.decode("utf-8", errors="ignore")
            reader = csv.DictReader(io.StringIO(text))
            return [dict(item) for item in reader]

        if extension == "xlsx":
            try:
                from openpyxl import load_workbook
            except Exception as exc:  # noqa: BLE001
                raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="openpyxl is required for xlsx uploads") from exc
            workbook = load_workbook(io.BytesIO(content), data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
            output: list[dict] = []
            for values in rows[1:]:
                row = {}
                for header, value in zip(headers, values):
                    if not header:
                        continue
                    row[header] = value
                output.append(row)
            return output

        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported topic bulk file type")

    def bulk_upsert_topics_from_file(self, *, file_name: str, content: bytes) -> TopicBulkUpsertResult:
        rows = self._parse_file_rows(file_name=file_name, content=content)
        parsed_rows: list[TopicBulkRow] = []
        parse_errors: list[ImportRowError] = []
        for index, raw in enumerate(rows, start=1):
            try:
                parsed_rows.append(TopicBulkRow.model_validate(raw))
            except Exception as exc:  # noqa: BLE001
                parse_errors.append(ImportRowError(row_number=index, errors=[str(exc)]))

        result = self.bulk_upsert_topics(TopicBulkUpsertRequest(rows=parsed_rows))
        if parse_errors:
            result.errors = result.errors + parse_errors
            result.skipped_count += len(parse_errors)
        return result
